import datetime
import os

import openpyxl
from dateutil.relativedelta import relativedelta
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import IntegrityError
from telegram_bot.models import EmployeeReport

User = get_user_model()

current_date = datetime.datetime.now()
last_month = current_date - relativedelta(months=1)

year_l = last_month.year
month_l = last_month.month
# Convert the numeric month to a string representation
# month_string = month_l.strftime('%B')

class Command(BaseCommand):

    def add_arguments(self, parser):
        # Add any command-specific arguments here
        parser.add_argument('file_name', type=str, help='Excel report file name')

    def handle(self, *args, **options):
        file_name = options['file_name']
        file_path = os.path.join(os.path.dirname(__file__) + f"/../../reports/{file_name}")

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        report_list = []



        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            # Reset the phone_number variable at the beginning of each iteration
            phone_number = None
            print(f"Populating--", row)
            id, id, user_name, phone_number, position, department, oclade_tarif, id, id, id, id, id, id, id, id, id, id, id, militar_regist, compensation_unused_vac, study, premium_9May, injury, premium_womenDay, premium_Navruz, premium_kurban, premium_constitution_retire, health_problem_one, lose_feeder_help_Xodjiyev, premium_independentDay, compensation_work_stop, compensation_household_products, prochi_nachisleniya, hospitalAUP, trudovoy_kodeks, premium_womenDay_retire, material_help_retire_injury, premium_constitution, prikaz_product, school_items, lose_feeder_help, otpusk_ligotniy, oclade_WTF, id, oclade_repairment_WTF, id, Saturday_work_WTF, id, tariff_WTF, id, Sdelno, hospital_WTF, id, vacation_WTF, id, vacation_add_WTF, id, vacation_pregnancy, night_shift_WTF, id, surcharge, surcharge_ragional_coef, surcharge_acc_ord, clasify, harmfulness, surcharge_advance_pay, loyalty, neyavka, cumulative_surcharge, material_help_pansion, material_help_marriage, material_help, material_help_cure, material_help_tough_situation, pansion, childcare_2_and_3, childcare_upto_2, travel, premium_ramadan, premium_one_time_racism, premium_anniversary, last_month_account, premium_house, premium_monthly, premium_executives, day_off, premium_award_diploma, labor_agreement, premium_holiday, premium_quarters, last_month_account_loyalty, surcharge_harm_repairment_WTF, id, vacation_contract, premium_order_advice, afghan_war_people, premium_PMM, repairment, per_day_limit, salary_13month, premium_travel, premium_general, per_day_full_limit, premium_motivation, maternity_leave_WTF, id, material_help_pansion_starting, nutrition_WTF, id, premium_skvajini, premium_contract, premium_svet, material_help_death, premium_TB, schooler, New_Year, region, material_help_vac, anniversaryx12, salary_13th, salary, communal_fee, vozvrat_debitor_fee, tax_free_education, book_take_fee, inventarization_fee, usherb_return_1, insurance_life, sog_zayav_Rakhimov, yuridik_uslugi, id, return_salary, zarabotnaya_plata, dogovor_dostovMizrob, accumulator_fee, tmz_fee, Sogspravki_fee, income_tax_manual, Gosposhlina_fee, telephone_fee, driver_licence_fee, avtouslugi_worker_fee, insurance, usherb_return_2, postal_expenditures, vozmesheni_salary, militar_regist_fee, JO_fee, contract_return, car_fee, micro_loan, id, recovery_fee, water_fee, dogovor_Buxoro_EKO_TUR, dogovor_5135LGM, jurnal_order3_fee, tax, vznos_pensiya, vznos_profsoyuz, reluctant_INPS, plastik_karta, id, id, alimony, alimony_postal_gain, uchebnoy_zavedeniya_fee, clasify_fee, viplata_otpusk, viplata_otpusk_dogovor, hostel_fee, vznos_partly, pereraschot, saturday_fee, campus_fee, sanatorium_vouchers_fee, fee_general, hotel_fee, bank_fee, phone_monthly_fee, GSM_fee, GorGaz_fee, land_owe_fee, nutrition_fee, electricity_fee, loan_credit_fee, loan_credit_ipoteka_no_ligota, loan_credit_OVERDRAFT, loan_credit_ipoteka, study_fee, loan_credit_consume, special_uniform_fee, sanksium_fee_charge_workers, sanksium_general, canteen_rent, report_fee, home_connect_fee, fine, id, id, remain, id, id, *_ = row
            
            
             # Evaluate formulas for the columns with Excel formulas
            # Skip if phone_number is empty
            if not phone_number:
                print(f"Skipping row {index + 2} as phone_number is empty")
                continue
            elif phone_number:
                # def evaluate_formulas(worksheet):
                #     worksheet = 'reports/oylik.xslx'
                #     for row in worksheet.iter_rows(values_only=True):
                #         for cell in row:
                #             if cell is not None and isinstance(cell, str) and cell.startswith("="):
                #                 cell.value = worksheet[cell.coordinate].value
                # # todo get value of remain from its excel formula
                # remain_val = sheet.cell(row=index + 1, column=6).value
                # remain_val = 0


                # Get first name and last name from the contact message
                name_parts = user_name.split(" ")
                first_name = name_parts[0]
                last_name = name_parts[1]
                middle_name = " ".join(name_parts[2:]) if len(name_parts) >= 4 else ""
                
                try:
                    user = User.objects.get(phone_number=str(phone_number))
                    if user:
                        # Update the user's first_name and last_name
                        user.first_name = first_name
                        user.last_name = last_name
                        user.middle_name = middle_name  # Set the middle_name
                        user.save()
                        created = False  # Set created to False since the user already existed
                    else:
                        # If the user doesn't exist, create a new one
                        user = User.objects.create(phone_number=str(phone_number), first_name=first_name, last_name=last_name, middle_name=middle_name)
                        created = True
                except IntegrityError as e:
                    # Handle the integrity error if needed
                    print(f"Error: {e}")
                except User.DoesNotExist:
                    # If the user doesn't exist, create a new one
                    user = User.objects.create(phone_number=str(phone_number), first_name=first_name, last_name=last_name, middle_name=middle_name)
                    created = True
                user, created = User.objects.get_or_create(phone_number=str(phone_number), first_name=first_name, last_name=last_name, middle_name=middle_name)

                print("User--", user.phone_number, "Created---", created)
                report = EmployeeReport(
                    user=user,
                    position=position,
                    department=department,
                    oclade_tarif=oclade_tarif,
                    militar_regist = militar_regist,
                    compensation_unused_vac = compensation_unused_vac,
                    study = study,
                    premium_9May = premium_9May,
                    injury = injury,
                    premium_womenDay = premium_womenDay,
                    premium_Navruz = premium_Navruz,
                    premium_kurban = premium_kurban,
                    premium_constitution_retire = premium_constitution_retire,
                    health_problem_one = health_problem_one,
                    lose_feeder_help_Xodjiyev = lose_feeder_help_Xodjiyev,
                    premium_independentDay = premium_independentDay,
                    compensation_work_stop = compensation_work_stop,
                    compensation_household_products = compensation_household_products,
                    prochi_nachisleniya = prochi_nachisleniya,
                    hospitalAUP = hospitalAUP,
                    trudovoy_kodeks = trudovoy_kodeks,
                    premium_womenDay_retire = premium_womenDay_retire,
                    material_help_retire_injury = material_help_retire_injury,
                    premium_constitution = premium_constitution,
                    prikaz_product = prikaz_product,
                    school_items = school_items,
                    lose_feeder_help = lose_feeder_help,
                    otpusk_ligotniy = otpusk_ligotniy,
                    oclade_WTF = oclade_WTF,
                    oclade_repairment_WTF = oclade_repairment_WTF,
                    Saturday_work_WTF = Saturday_work_WTF,
                    tariff_WTF = tariff_WTF,
                    Sdelno = Sdelno,
                    hospital_WTF = hospital_WTF,
                    vacation_WTF = vacation_WTF,
                    vacation_add_WTF = vacation_add_WTF,
                    vacation_pregnancy = vacation_pregnancy,
                    night_shift_WTF = night_shift_WTF,
                    surcharge = surcharge,
                    surcharge_ragional_coef = surcharge_ragional_coef,
                    surcharge_acc_ord = surcharge_acc_ord,
                    clasify = clasify,
                    harmfulness = harmfulness,
                    surcharge_advance_pay = surcharge_advance_pay,
                    loyalty = loyalty,
                    neyavka = neyavka,
                    cumulative_surcharge = cumulative_surcharge,
                    material_help_pansion = material_help_pansion,
                    material_help_marriage = material_help_marriage,
                    material_help = material_help,
                    material_help_cure = material_help_cure,
                    material_help_tough_situation = material_help_tough_situation,
                    pansion = pansion,
                    childcare_2_and_3 = childcare_2_and_3,
                    childcare_upto_2 = childcare_upto_2,
                    travel = travel,
                    premium_ramadan = premium_ramadan,
                    premium_one_time_racism = premium_one_time_racism,
                    premium_anniversary = premium_anniversary,
                    last_month_account = last_month_account,
                    premium_house = premium_house,
                    premium_monthly = premium_monthly,
                    premium_executives = premium_executives,
                    day_off = day_off,
                    premium_award_diploma = premium_award_diploma,
                    labor_agreement = labor_agreement,
                    premium_holiday = premium_holiday,
                    premium_quarters = premium_quarters,
                    last_month_account_loyalty = last_month_account_loyalty,
                    surcharge_harm_repairment_WTF = surcharge_harm_repairment_WTF,
                    vacation_contract = vacation_contract,
                    premium_order_advice = premium_order_advice,
                    afghan_war_people = afghan_war_people,
                    premium_PMM = premium_PMM,
                    repairment = repairment,
                    per_day_limit = per_day_limit,
                    salary_13month = salary_13month,
                    premium_travel = premium_travel,
                    premium_general = premium_general,
                    per_day_full_limit = per_day_full_limit,
                    premium_motivation = premium_motivation,
                    maternity_leave_WTF = maternity_leave_WTF,
                    material_help_pansion_starting = material_help_pansion_starting,
                    nutrition_WTF = nutrition_WTF,
                    premium_skvajini = premium_skvajini,
                    premium_contract = premium_contract,
                    premium_svet = premium_svet,
                    material_help_death = material_help_death,
                    premium_TB = premium_TB,
                    schooler = schooler,
                    New_Year = New_Year,
                    region = region,
                    material_help_vac = material_help_vac,
                    anniversaryx12 = anniversaryx12,
                    salary_13th = salary_13th,
                    salary = salary,
                    communal_fee = communal_fee,
                    vozvrat_debitor_fee = vozvrat_debitor_fee,
                    tax_free_education = tax_free_education,
                    book_take_fee = book_take_fee,
                    inventarization_fee = inventarization_fee,
                    usherb_return_1 = usherb_return_1,
                    insurance_life = insurance_life,
                    sog_zayav_Rakhimov = sog_zayav_Rakhimov,
                    yuridik_uslugi = yuridik_uslugi,
                    return_salary = return_salary,
                    zarabotnaya_plata = zarabotnaya_plata,
                    dogovor_dostovMizrob = dogovor_dostovMizrob,
                    accumulator_fee = accumulator_fee,
                    tmz_fee = tmz_fee,
                    Sogspravki_fee = Sogspravki_fee,
                    income_tax_manual = income_tax_manual,
                    Gosposhlina_fee = Gosposhlina_fee,
                    telephone_fee = telephone_fee,
                    driver_licence_fee = driver_licence_fee,
                    avtouslugi_worker_fee = avtouslugi_worker_fee,
                    insurance = insurance,
                    usherb_return_2 = usherb_return_2,
                    postal_expenditures = postal_expenditures,
                    vozmesheni_salary = vozmesheni_salary,
                    militar_regist_fee = militar_regist_fee,
                    JO_fee = JO_fee,
                    contract_return = contract_return,
                    car_fee = car_fee,
                    micro_loan = micro_loan,
                    recovery_fee = recovery_fee,
                    water_fee = water_fee,
                    dogovor_Buxoro_EKO_TUR = dogovor_Buxoro_EKO_TUR,
                    dogovor_5135LGM = dogovor_5135LGM,
                    jurnal_order3_fee = jurnal_order3_fee,
                    tax = tax,
                    vznos_pensiya = vznos_pensiya,
                    vznos_profsoyuz = vznos_profsoyuz,
                    reluctant_INPS = reluctant_INPS,
                    plastik_karta = plastik_karta,
                    alimony = alimony,
                    alimony_postal_gain = alimony_postal_gain,
                    uchebnoy_zavedeniya_fee = uchebnoy_zavedeniya_fee,
                    clasify_fee = clasify_fee,
                    viplata_otpusk = viplata_otpusk,
                    viplata_otpusk_dogovor = viplata_otpusk_dogovor,
                    hostel_fee = hostel_fee,
                    vznos_partly = vznos_partly,
                    pereraschot = pereraschot,
                    saturday_fee = saturday_fee,
                    campus_fee = campus_fee,
                    sanatorium_vouchers_fee = sanatorium_vouchers_fee,
                    fee_general = fee_general,
                    hotel_fee = hotel_fee,
                    bank_fee = bank_fee,
                    phone_monthly_fee = phone_monthly_fee,
                    GSM_fee = GSM_fee,
                    GorGaz_fee = GorGaz_fee,
                    land_owe_fee = land_owe_fee,
                    nutrition_fee = nutrition_fee,
                    electricity_fee = electricity_fee,
                    loan_credit_fee = loan_credit_fee,
                    loan_credit_ipoteka_no_ligota = loan_credit_ipoteka_no_ligota,
                    loan_credit_OVERDRAFT = loan_credit_OVERDRAFT,
                    loan_credit_ipoteka = loan_credit_ipoteka,
                    study_fee = study_fee,
                    loan_credit_consume = loan_credit_consume,
                    special_uniform_fee = special_uniform_fee,
                    sanksium_fee_charge_workers = sanksium_fee_charge_workers,
                    sanksium_general = sanksium_general,
                    canteen_rent = canteen_rent,
                    report_fee = report_fee,
                    home_connect_fee = home_connect_fee,
                    fine = fine,
                    remain = remain,
                    year = year_l,
                    month = month_l,

                )
                report_list.append(report)
            else:
                print(f"phone number is required. Now it is - {phone_number}")
        print(f"Populating {len(report_list)} reports")
        EmployeeReport.objects.bulk_create(report_list)


